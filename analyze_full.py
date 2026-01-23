# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
import os

output_file = r"c:\Users\lyen\Downloads\EU Quota\analysis_full.txt"
template_file = r"c:\Users\lyen\Downloads\EU Quota\MEPS European Steel Review Quota Update - Dec 2025.xlsx"
xlsm_file = r"c:\Users\lyen\Downloads\MEPS European Steel Review Quota Update - Dec 2025.xlsm"
urls_file = r"c:\Users\lyen\Downloads\EU Quota\EU Quota URL's.xlsx"

with open(output_file, 'w', encoding='utf-8') as f:
    # Analyze EU Quota URL's file
    f.write("=== EU QUOTA URL's FILE ===\n")
    try:
        wb_urls = load_workbook(urls_file)
        f.write(f"Sheets: {wb_urls.sheetnames}\n")
        for sn in wb_urls.sheetnames:
            df = pd.read_excel(urls_file, sheet_name=sn, header=None)
            f.write(f"\nSheet '{sn}': shape={df.shape}\n")
            # Print first 10 rows
            f.write("First 10 rows:\n")
            f.write(df.head(10).to_string() + "\n")
    except Exception as e:
        f.write(f"Error: {e}\n")
    
    f.write("\n\n=== TEMPLATE FILE - UK SHEET DETAILED ===\n")
    wb = load_workbook(template_file)
    ws_uk = wb['United Kingdom']
    
    # Print all rows up to 30 to see full UK data
    f.write("First 30 rows of UK sheet:\n")
    for row in range(1, min(31, ws_uk.max_row + 1)):
        row_data = []
        for col in range(1, ws_uk.max_column + 1):
            cell = ws_uk.cell(row, col)
            val = str(cell.value)[:40] if cell.value else ""
            row_data.append(val)
        f.write(f"Row {row:2d}: {row_data}\n")
    
    # Get all unique categories in UK sheet
    f.write("\n\nUnique values in UK sheet Column A (Quota Category) from row 16 onwards:\n")
    categories = set()
    for row in range(16, ws_uk.max_row + 1):
        val = ws_uk.cell(row, 1).value
        if val:
            categories.add(val)
    for cat in sorted(categories):
        f.write(f"  - {cat}\n")
    
    # Get all unique countries in UK sheet
    f.write("\nUnique values in UK sheet Column B (Country) from row 16 onwards:\n")
    countries = set()
    for row in range(16, ws_uk.max_row + 1):
        val = ws_uk.cell(row, 2).value
        if val:
            countries.add(val)
    for country in sorted(countries):
        f.write(f"  - {country}\n")
    
    # Check XLSM file
    f.write("\n\n=== XLSM FILE (MACRO-ENABLED) ===\n")
    try:
        if os.path.exists(xlsm_file):
            wb_xlsm = load_workbook(xlsm_file, keep_vba=True)
            f.write(f"Sheets: {wb_xlsm.sheetnames}\n")
            for sn in wb_xlsm.sheetnames:
                ws = wb_xlsm[sn]
                f.write(f"\nSheet '{sn}': max_row={ws.max_row}, max_col={ws.max_column}\n")
                # Check first few rows
                for row in range(1, min(5, ws.max_row + 1)):
                    row_data = [ws.cell(row, col).value for col in range(1, min(ws.max_column + 1, 8))]
                    f.write(f"  Row {row}: {row_data}\n")
        else:
            f.write(f"File not found: {xlsm_file}\n")
    except Exception as e:
        f.write(f"Error reading XLSM: {e}\n")
    
    # Check the EU sheet too for categories comparison
    ws_eu = wb['European Union']
    f.write("\n\n=== EU SHEET - Unique Categories ===\n")
    eu_categories = set()
    for row in range(16, ws_eu.max_row + 1):
        val = ws_eu.cell(row, 1).value
        if val:
            eu_categories.add(val)
    for cat in sorted(eu_categories):
        f.write(f"  - {cat}\n")

print("Full analysis saved to:", output_file)
