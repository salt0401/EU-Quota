# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border
import os

output_file = r"c:\Users\lyen\Downloads\EU Quota\analysis_detailed.txt"
template_file = r"c:\Users\lyen\Downloads\EU Quota\MEPS European Steel Review Quota Update - Dec 2025.xlsx"

with open(output_file, 'w', encoding='utf-8') as f:
    wb = load_workbook(template_file)
    
    # Focus on European Union sheet structure
    ws = wb['European Union']
    f.write("=== EUROPEAN UNION SHEET DETAILED ANALYSIS ===\n\n")
    
    # Print first 20 rows with all columns to understand the structure
    f.write("First 25 rows, all columns:\n")
    for row in range(1, 26):
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            val = str(cell.value) if cell.value else ""
            row_data.append(val[:50])  # truncate long values
        f.write(f"Row {row:2d}: {row_data}\n")
    
    f.write("\n\n=== DATA TABLE HEADER ROW ===\n")
    # Find the actual data table header row
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row, 1).value
        if cell_val and ('Category' in str(cell_val) or 'Product' in str(cell_val)):
            f.write(f"Found data header at row {row}\n")
            header_row = row
            headers = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
            f.write(f"Headers: {headers}\n")
            # Get formatting info
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if cell.value:
                    f.write(f"  Col {col} '{cell.value}': ")
                    f.write(f"font_bold={cell.font.bold}, font_size={cell.font.size}, ")
                    if cell.fill.start_color.type == 'rgb':
                        f.write(f"fill={cell.fill.start_color.rgb}, ")
                    f.write(f"align_h={cell.alignment.horizontal}\n")
            break
    
    f.write("\n\n=== SAMPLE DATA ROWS AFTER HEADER ===\n")
    # Print 5 data rows after the header
    for row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
        row_data = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        f.write(f"Row {row}: {row_data}\n")
    
    f.write("\n\n=== CELL FORMATTING DETAILS ===\n")
    # Check formatting of specific cells
    for row in [1, 2, 3, 4, header_row, header_row+1]:
        for col in range(1, min(ws.max_column + 1, 8)):
            cell = ws.cell(row, col)
            if cell.value:
                f.write(f"Cell ({row},{col}) '{str(cell.value)[:30]}...':\n")
                f.write(f"  Font: bold={cell.font.bold}, size={cell.font.size}, color={cell.font.color.rgb if cell.font.color and cell.font.color.type=='rgb' else 'default'}\n")
                f.write(f"  Fill: {cell.fill.start_color.rgb if cell.fill.start_color.type=='rgb' else cell.fill.patternType}\n")
                f.write(f"  Align: h={cell.alignment.horizontal}, v={cell.alignment.vertical}\n")
                if cell.border.left.style:
                    f.write(f"  Border: left={cell.border.left.style}\n")
    
    f.write("\n\n=== UNITED KINGDOM SHEET STRUCTURE ===\n")
    ws_uk = wb['United Kingdom']
    
    # Print first 20 rows
    f.write("First 25 rows, all columns:\n")
    for row in range(1, 26):
        row_data = []
        for col in range(1, ws_uk.max_column + 1):
            cell = ws_uk.cell(row, col)
            val = str(cell.value) if cell.value else ""
            row_data.append(val[:50])
        f.write(f"Row {row:2d}: {row_data}\n")

print("Detailed analysis saved to:", output_file)
