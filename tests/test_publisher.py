# -*- coding: utf-8 -*-
"""
Tests for src/publisher.py - published-data writer for the daily Actions run
"""
import csv
import io
import json
import os
import sys

import pandas as pd
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.publisher import (
    HISTORY_COLUMNS,
    build_eu_history_rows,
    build_uk_history_rows,
    update_history_csv,
    generate_history_xlsx,
    publish_data,
)


def _eu_df():
    # internal EU metrics: kg volumes, 0-100 percentages
    return pd.DataFrame({
        'input_order_number': ['099801', '099600'],
        'input_quota_category': ['Hot Rolled - 1.A', 'Hot Rolled - 1.A'],
        'input_country': ['Türkiye', 'Other countries'],
        'origin': ['European Union Türkiye', 'ERGA OMNES'],
        'quota_limit': [160573740.0, 5564010.0],
        'quota_allocated': [160573740.0, 0.0],
        'balance_remaining': [0.0, 5564010.0],
        'pct_allocated': [100.0, 0.0],
        'pct_remaining': [0.0, 100.0],
    })


def _uk_df():
    # internal UK metrics: tonne volumes, 0-1 percentages
    return pd.DataFrame({
        'input_order_number': ['058600'],
        'input_quota_category': ['Hot rolled - 1'],
        'input_country': ['European Union'],
        'quota_limit_tonnes': [93750.0],
        'quota_allocated_tonnes': [3501.5],
        'balance_remaining_tonnes': [90248.5],
        'pct_allocated': [0.037349],
        'pct_remaining': [0.962651],
    })


class TestHistoryRows:

    def test_eu_rows_converted_to_tonnes(self):
        rows = build_eu_history_rows(_eu_df(), '2026-07-06')
        assert len(rows) == 2
        assert rows[0]['quota_limit_t'] == 160573.74
        assert rows[0]['region'] == 'EU'
        assert rows[0]['country'] == 'Türkiye'  # curated name, not origin
        assert rows[0]['pct_allocated'] == 100.0
        assert rows[0]['scrape_status'] == 'ok'

    def test_uk_rows_pct_scaled_to_0_100(self):
        rows = build_uk_history_rows(_uk_df(), '2026-07-06')
        assert len(rows) == 1
        assert rows[0]['region'] == 'UK'
        assert rows[0]['quota_limit_t'] == 93750.0
        assert rows[0]['pct_allocated'] == pytest.approx(3.73, abs=0.01)

    def test_failed_status_carried_through(self):
        df = _eu_df()
        df['scrape_status'] = [None, 'failed']
        rows = build_eu_history_rows(df, '2026-07-06')
        assert rows[0]['scrape_status'] == 'ok'
        assert rows[1]['scrape_status'] == 'failed'

    def test_empty_frames(self):
        assert build_eu_history_rows(pd.DataFrame(), '2026-07-06') == []
        assert build_uk_history_rows(None, '2026-07-06') == []


class TestHistoryCsv:

    def test_creates_and_appends(self, tmp_path):
        path = str(tmp_path / 'quota_history.csv')
        day1 = build_eu_history_rows(_eu_df(), '2026-07-06')
        assert update_history_csv(path, day1, '2026-07-06') == 2
        day2 = build_eu_history_rows(_eu_df(), '2026-07-07')
        assert update_history_csv(path, day2, '2026-07-07') == 4

        with io.open(path, encoding='utf-8-sig', newline='') as f:
            rows = list(csv.DictReader(f))
        assert [r['date'] for r in rows] == ['2026-07-06', '2026-07-06',
                                             '2026-07-07', '2026-07-07']

    def test_rerun_same_date_is_idempotent(self, tmp_path):
        path = str(tmp_path / 'quota_history.csv')
        rows = build_eu_history_rows(_eu_df(), '2026-07-06')
        update_history_csv(path, rows, '2026-07-06')
        total = update_history_csv(path, rows, '2026-07-06')  # re-run same day
        assert total == 2  # replaced, not duplicated

    def test_single_region_rerun_preserves_other_region(self, tmp_path):
        # A same-date re-run that produced only EU rows must NOT delete the
        # date's existing UK rows (e.g. workflow retry with a broken UK input)
        path = str(tmp_path / 'quota_history.csv')
        both = (build_eu_history_rows(_eu_df(), '2026-07-06')
                + build_uk_history_rows(_uk_df(), '2026-07-06'))
        assert update_history_csv(path, both, '2026-07-06') == 3
        eu_only = build_eu_history_rows(_eu_df(), '2026-07-06')
        total = update_history_csv(path, eu_only, '2026-07-06')
        assert total == 3  # UK row survived
        with io.open(path, encoding='utf-8-sig', newline='') as f:
            regions = {r['region'] for r in csv.DictReader(f)}
        assert regions == {'EU', 'UK'}

    def test_unicode_country_survives_roundtrip(self, tmp_path):
        path = str(tmp_path / 'quota_history.csv')
        update_history_csv(path, build_eu_history_rows(_eu_df(), '2026-07-06'),
                           '2026-07-06')
        with io.open(path, encoding='utf-8-sig', newline='') as f:
            rows = {r['order_number']: r for r in csv.DictReader(f)}
        assert rows['099801']['country'] == 'Türkiye'


class TestPublishData:

    def test_full_publish(self, tmp_path):
        report = tmp_path / 'report.xlsx'
        report.write_bytes(b'PK\x03\x04fake')
        publish_dir = str(tmp_path / 'published')

        from datetime import date
        meta = publish_data(_eu_df(), _uk_df(), str(report), publish_dir,
                            run_date=date(2026, 7, 6),
                            period_display='01-Jul-2026 to 30-Sep-2026')

        assert os.path.exists(os.path.join(publish_dir, 'MEPS_Quota_Update_latest.xlsx'))
        assert os.path.exists(os.path.join(publish_dir, 'quota_history.csv'))
        assert os.path.exists(os.path.join(publish_dir, 'Quota_History.xlsx'))
        assert meta['eu_quotas'] == 2
        assert meta['uk_quotas'] == 1
        assert meta['history_rows'] == 3
        assert meta['data_date'] == '2026-07-06'

        with io.open(os.path.join(publish_dir, 'metadata.json'), encoding='utf-8') as f:
            on_disk = json.load(f)
        assert on_disk['quota_period'] == '01-Jul-2026 to 30-Sep-2026'

    def test_refuses_mostly_failed_scrape(self, tmp_path):
        # >50% failures means the source is broken (or the regulation was
        # renewed with new order numbers) — publishing would serve garbage
        from datetime import date
        report = tmp_path / 'report.xlsx'
        report.write_bytes(b'PK\x03\x04fake')
        df = _eu_df()
        df['scrape_status'] = ['failed', 'failed']
        with pytest.raises(RuntimeError, match='failed to scrape'):
            publish_data(df, _uk_df(), str(report), str(tmp_path / 'pub'),
                         run_date=date(2026, 7, 6))

    def test_refuses_expired_quota_window(self, tmp_path):
        # Stale StartDate: TARIC returns the frozen state of an expired
        # window as a successful scrape — the gate must catch it
        from datetime import date
        report = tmp_path / 'report.xlsx'
        report.write_bytes(b'PK\x03\x04fake')
        df = _eu_df()
        df['validity_end'] = ['30-09-2026', '30-09-2026']
        with pytest.raises(RuntimeError, match='stale'):
            publish_data(df, _uk_df(), str(report), str(tmp_path / 'pub'),
                         run_date=date(2026, 10, 2))
        # ...and passes while the window is still open
        meta = publish_data(df, _uk_df(), str(report), str(tmp_path / 'pub2'),
                            run_date=date(2026, 7, 6))
        assert meta['eu_quotas'] == 2

    def test_history_xlsx_has_region_sheets(self, tmp_path):
        path = str(tmp_path / 'quota_history.csv')
        rows = (build_eu_history_rows(_eu_df(), '2026-07-06')
                + build_uk_history_rows(_uk_df(), '2026-07-06'))
        update_history_csv(path, rows, '2026-07-06')
        xlsx = str(tmp_path / 'Quota_History.xlsx')
        generate_history_xlsx(path, xlsx)

        import openpyxl
        wb = openpyxl.load_workbook(xlsx)
        assert wb.sheetnames == ['EU History', 'UK History']
        eu = wb['EU History']
        assert [c.value for c in eu[1]] == HISTORY_COLUMNS
        assert eu.max_row == 3  # header + 2 EU rows
        assert wb['UK History'].max_row == 2
        # rows are sorted by order number: row 2 = 099600, row 3 = 099801
        assert eu.cell(row=3, column=6).value == 160573.74  # numeric, not text
