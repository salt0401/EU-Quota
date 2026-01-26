# -*- coding: utf-8 -*-
"""
Create COMPLETE UK input file with ALL 71 rows from template
Each row maps to an order number, even if multiple rows share the same order
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Complete UK Template Data (all 71 rows extracted from meps_customer_template.xlsx)
# Format: (Category, Country, Quota_Limit_Tonnes)
UK_TEMPLATE_ROWS = [
    ("Non-alloy and other alloy hot-rolled sheet and strip 1A", "European Union", 187671),
    ("Non-alloy and other alloy hot-rolled sheet and strip 1A", "Turkey", 24641),
    ("Non-alloy and other alloy hot-rolled sheet and strip 1A", "Taiwan", 13718),
    ("Non-alloy and other alloy hot-rolled sheet and strip 1A", "All others", 23611),
    ("Non-alloy and other alloy hot-rolled sheet and strip 1B", "All others", 579165),
    ("Metallic coated sheet", "European Union", 324154),
    ("Metallic coated sheet", "Taiwan", 33513),
    ("Metallic coated sheet", "India", 24752),
    ("Metallic coated sheet", "All others", 85595),
    ("Metallic coated sheet", "Japan*", 12839),
    ("Metallic coated sheet", "South Korea*", 12839),
    ("Metallic coated sheet", "Taiwan*", 12839),
    ("Metallic coated sheet", "Algeria*", 12839),
    ("Metallic coated sheet", "Singapore*", 12839),
    ("Metallic coated sheet", "United States*", 12839),
    ("Organic coated sheet", "European Union", 36684),
    ("Organic coated sheet", "Republic of Korea", 14907),
    ("Organic coated sheet", "All others", 2228),
    ("Tin Mill products", "European Union", 31986),
    ("Tin Mill products", "China", 8121),
    ("Tin Mill products", "Republic of Korea", 2518),
    ("Tin Mill products", "Taiwan", 2652),
    ("Tin Mill products", "All others", 1085),
    ("Non-alloy and Other Alloy Quarto Plates", "European Union", 71178),
    ("Non-alloy and Other Alloy Quarto Plates", "All others", 25401),
    ("Non-alloy and Other Alloy Quarto Plates", "South Korea*", 5080),
    ("Non-alloy and Other Alloy Quarto Plates", "Taiwan*", 5080),
    ("Non-alloy and Other Alloy Quarto Plates", "Algeria*", 5080),
    ("Non-alloy and Other Alloy Quarto Plates", "United States*", 5080),
    ("Non-alloy and Other Alloy Quarto Plates", "Australia*", 5080),
    ("Non-alloy and Other Alloy Quarto Plates", "Canada*", 5080),
    ("Alloy merchant bars and light sections", "European Union", 29517),
    ("Alloy merchant bars and light sections", "All others", 4246),
    ("Non-alloy merchant bars and light sections", "European Union", 35458),
    ("Non-alloy merchant bars and light sections", "Turkey", 13346),
    ("Non-alloy merchant bars and light sections", "All others", 7585),
    ("Rebars", "European Union", 75521),
    ("Rebars", "Turkey", 35304),
    ("Rebars", "All others", 24037),
    ("Rebars", "Egypt*", 4807),
    ("Rebars", "Vietnam*", 4807),
    ("Rebars", "Algeria*", 4807),
    ("Rebars", "New Zealand*", 4807),
    ("Rebars", "Norway*", 4807),
    ("Rebars", "Taiwan*", 4807),
    ("Non-alloy and Other Alloy Wire Rod", "European Union", 75021),
    ("Non-alloy and Other Alloy Wire Rod", "All others", 3275),
    ("Angles, Shapes and Sections of Iron or Non-alloy Steel", "European Union", 170816),
    ("Angles, Shapes and Sections of Iron or Non-alloy Steel", "All others", 17753),
    ("Gas pipes", "European Union", 4797),
    ("Gas pipes", "All others", 141),
    ("Hollow sections", "European Union", 7079),
    ("Hollow sections", "India", 3640),
    ("Hollow sections", "Turkey", 15728),
    ("Hollow sections", "All others", 740),
    ("Large welded tubes - 25A", "European Union", 11332),
    ("Large welded tubes - 25A", "Turkey", 37335),
    ("Large welded tubes - 25A", "All others", 3445),
    ("Large welded tubes - 25A", "European Union", 6333),  # Row 74 - different EU allocation?
    ("Large welded tubes - 25A", "Japan", 8276),
    ("Large welded tubes - 25A", "Republic of Korea", 1265),
    ("Large welded tubes - 25A", "All others", 2232),
    ("Large welded tubes - 25B", "All others", 16290),
    ("Large welded tubes - 25B", "European Union", 2051),
    ("Large welded tubes - 25B", "Japan", 4682),
    ("Large welded tubes - 25B", "Republic of Korea", 4938),
    ("Other Welded Pipes", "All others", 22820),
    ("Other Welded Pipes", "European Union", 5860),
    ("Other Welded Pipes", "China", 11140),
    ("Other Welded Pipes", "Turkey", 15335),
    ("Other Welded Pipes", "United Arab Emirates", 10118),
]

# Category code mapping
CATEGORY_CODES = {
    "Non-alloy and other alloy hot-rolled sheet and strip 1A": "1A",
    "Non-alloy and other alloy hot-rolled sheet and strip 1B": "1B",
    "Metallic coated sheet": "4",
    "Organic coated sheet": "5",
    "Tin Mill products": "6",
    "Non-alloy and Other Alloy Quarto Plates": "7",
    "Alloy merchant bars and light sections": "12A",
    "Non-alloy merchant bars and light sections": "12B",
    "Rebars": "13",
    "Non-alloy and Other Alloy Wire Rod": "16",
    "Angles, Shapes and Sections of Iron or Non-alloy Steel": "17",
    "Gas pipes": "20",
    "Hollow sections": "21",
    "Large welded tubes - 25A": "25A",
    "Large welded tubes - 25B": "25B",
    "Other Welded Pipes": "26",
}

# Country to order code mapping
COUNTRY_CODES = {
    "European Union": "EU",
    "Turkey": "Turkey",
    "Taiwan": "Taiwan",
    "India": "India",
    "China": "China",
    "Japan": "Japan",
    "Republic of Korea": "South_Korea",
    "United Arab Emirates": "UAE",
    "All others": "All_others",
    # Countries with * map to All_others order number
    "Japan*": "All_others",
    "South Korea*": "All_others",
    "Taiwan*": "All_others",
    "Algeria*": "All_others",
    "Singapore*": "All_others",
    "United States*": "All_others",
    "Egypt*": "All_others",
    "Vietnam*": "All_others",
    "Australia*": "All_others",
    "Canada*": "All_others",
    "New Zealand*": "All_others",
    "Norway*": "All_others",
}

# Order numbers from Trade Remedies Notice 2025/12
ORDER_NUMBERS = {
    ("1A", "EU"): "058001",
    ("1A", "Turkey"): "058967",
    ("1A", "Taiwan"): "058085",
    ("1A", "All_others"): "058002",
    ("1B", "All_others"): "058003",
    ("4", "EU"): "058006",
    ("4", "Taiwan"): "058088",
    ("4", "India"): "058106",
    ("4", "All_others"): "058007",
    ("5", "EU"): "058010",
    ("5", "South_Korea"): "058827",
    ("5", "All_others"): "058011",
    ("6", "EU"): "058012",
    ("6", "China"): "058831",
    ("6", "South_Korea"): "058097",
    ("6", "Taiwan"): "058098",
    ("6", "All_others"): "058013",
    ("7", "EU"): "058014",
    ("7", "All_others"): "058015",
    ("12A", "EU"): "058100",
    ("12A", "All_others"): "058102",
    ("12B", "EU"): "058103",
    ("12B", "Turkey"): "058104",
    ("12B", "All_others"): "058105",
    ("13", "EU"): "058018",
    ("13", "Turkey"): "058866",
    ("13", "All_others"): "058019",
    ("16", "EU"): "058026",
    ("16", "All_others"): "058027",
    ("17", "EU"): "058028",
    ("17", "All_others"): "058029",
    ("20", "EU"): "058032",
    ("20", "Turkey"): "058911",
    ("20", "India"): "058912",
    ("20", "All_others"): "058033",
    ("21", "EU"): "058034",
    ("21", "Turkey"): "058916",
    ("21", "India"): "058912",  # Same as Gas pipes India
    ("21", "All_others"): "058035",
    ("25A", "EU"): "058091",
    ("25A", "South_Korea"): "058095",
    ("25A", "Japan"): "058108",
    ("25A", "Turkey"): "058967",  # Need to verify
    ("25A", "All_others"): "058036",
    ("25B", "EU"): "058037",
    ("25B", "South_Korea"): "058974",
    ("25B", "Japan"): "058109",
    ("25B", "All_others"): "058038",
    ("26", "EU"): "058039",
    ("26", "UAE"): "058948",
    ("26", "Turkey"): "058947",
    ("26", "China"): "058949",
    ("26", "All_others"): "058041",
}

UK_URL_BASE = "https://www.trade-tariff.service.gov.uk/quota_search?order_number="


def get_order_number(category, country):
    """Get order number for a category/country combination"""
    cat_code = CATEGORY_CODES.get(category, "")
    country_code = COUNTRY_CODES.get(country, "")

    if not cat_code:
        return None

    key = (cat_code, country_code)
    return ORDER_NUMBERS.get(key)


def create_uk_input_file():
    """Create UK input file with all 71 rows"""

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_path = os.path.join(base_dir, "data", "input", "uk_quota_urls.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "UK Quota URLs"

    # Headers
    headers = ['Order Number', 'Quota Category', 'Country', 'Template Quota Limit', 'URL']

    # Style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Write data rows
    missing = []
    for row_idx, (category, country, quota_limit) in enumerate(UK_TEMPLATE_ROWS, 2):
        order_number = get_order_number(category, country)

        if not order_number:
            missing.append(f"Row {row_idx}: {category} | {country}")
            order_number = "UNKNOWN"

        ws.cell(row_idx, 1, order_number)
        ws.cell(row_idx, 2, category)
        ws.cell(row_idx, 3, country)
        ws.cell(row_idx, 4, quota_limit)

        # URL formula
        if order_number != "UNKNOWN":
            ws.cell(row_idx, 5, f'=CONCATENATE("{UK_URL_BASE}",A{row_idx})')
        else:
            ws.cell(row_idx, 5, "")

        # Apply borders
        for col in range(1, 6):
            ws.cell(row_idx, col).border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 70

    wb.save(output_path)
    wb.close()

    print(f"UK input file created: {output_path}")
    print(f"Total rows: {len(UK_TEMPLATE_ROWS)}")

    if missing:
        print(f"\nMISSING ORDER NUMBERS ({len(missing)}):")
        for m in missing:
            print(f"  ! {m}")

    # Print unique order numbers
    unique_orders = set()
    for category, country, _ in UK_TEMPLATE_ROWS:
        order = get_order_number(category, country)
        if order:
            unique_orders.add(order)

    print(f"\nUnique order numbers to scrape: {len(unique_orders)}")

    return output_path


if __name__ == "__main__":
    create_uk_input_file()
