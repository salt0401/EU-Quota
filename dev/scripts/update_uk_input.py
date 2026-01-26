# -*- coding: utf-8 -*-
"""
Update UK Quota Input File
Creates uk_quota_urls.xlsx with same format as EU quota_urls.xlsx
Updated for 2026 Q1 with correct order numbers

Format matches EU file exactly:
- Row 1: Base URL in column B
- Row 5: Headers
- Row 6+: Data with URL formula
- Excel Table for structured references
"""

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# UK quota data - updated for 2026 Q1
# Format: (order_number, category, country, quarter, template_quota_limit_tonnes)
uk_quotas = [
    # Category 1A - Non-alloy hot-rolled sheet
    (58001, "Non-alloy and other alloy hot-rolled sheet and strip 1A", "European Union", "2026-01-01", 187671),
    (58967, "Non-alloy and other alloy hot-rolled sheet and strip 1A", "Turkey", "2026-01-01", 24641),
    (58085, "Non-alloy and other alloy hot-rolled sheet and strip 1A", "Taiwan", "2026-01-01", 13718),
    (58002, "Non-alloy and other alloy hot-rolled sheet and strip 1A", "All others", "2026-01-01", 23611),

    # Category 1B - Non-alloy hot-rolled sheet (NEW structure for 2026)
    # Old 058003 (579,165 tonnes) replaced with 058110, 058111, 058112
    (58110, "Non-alloy and other alloy hot-rolled sheet and strip 1B", "All others (Sub-quota 1)", "2026-01-01", 113315),
    (58111, "Non-alloy and other alloy hot-rolled sheet and strip 1B", "All others (Sub-quota 2)", "2026-01-01", 226631),
    (58112, "Non-alloy and other alloy hot-rolled sheet and strip 1B", "All others (Sub-quota 3)", "2026-01-01", 193644),

    # Category 4 - Metallic coated sheet
    (58006, "Metallic coated sheet", "European Union", "2026-01-01", 324154),
    (58088, "Metallic coated sheet", "Taiwan", "2026-01-01", 33513),
    (58106, "Metallic coated sheet", "India", "2026-01-01", 24752),
    (58007, "Metallic coated sheet", "All others", "2026-01-01", 85595),
    (58007, "Metallic coated sheet", "Japan*", "2026-01-01", 12839),
    (58007, "Metallic coated sheet", "South Korea*", "2026-01-01", 12839),
    (58007, "Metallic coated sheet", "Taiwan*", "2026-01-01", 12839),
    (58007, "Metallic coated sheet", "Algeria*", "2026-01-01", 12839),
    (58007, "Metallic coated sheet", "Singapore*", "2026-01-01", 12839),
    (58007, "Metallic coated sheet", "United States*", "2026-01-01", 12839),

    # Category 5 - Organic coated sheet
    (58010, "Organic coated sheet", "European Union", "2026-01-01", 36684),
    (58827, "Organic coated sheet", "Republic of Korea", "2026-01-01", 14907),
    (58011, "Organic coated sheet", "All others", "2026-01-01", 2228),

    # Category 6 - Tin Mill products
    (58012, "Tin Mill products", "European Union", "2026-01-01", 31986),
    (58831, "Tin Mill products", "China", "2026-01-01", 8121),
    (58097, "Tin Mill products", "Republic of Korea", "2026-01-01", 2518),
    (58098, "Tin Mill products", "Taiwan", "2026-01-01", 2652),
    (58013, "Tin Mill products", "All others", "2026-01-01", 1085),

    # Category 7 - Non-alloy and Other Alloy Quarto Plates
    (58014, "Non-alloy and Other Alloy Quarto Plates", "European Union", "2026-01-01", 71178),
    (58015, "Non-alloy and Other Alloy Quarto Plates", "All others", "2026-01-01", 25401),
    (58015, "Non-alloy and Other Alloy Quarto Plates", "South Korea*", "2026-01-01", 5080),
    (58015, "Non-alloy and Other Alloy Quarto Plates", "Taiwan*", "2026-01-01", 5080),
    (58015, "Non-alloy and Other Alloy Quarto Plates", "Algeria*", "2026-01-01", 5080),
    (58015, "Non-alloy and Other Alloy Quarto Plates", "United States*", "2026-01-01", 5080),
    (58015, "Non-alloy and Other Alloy Quarto Plates", "Australia*", "2026-01-01", 5080),
    (58015, "Non-alloy and Other Alloy Quarto Plates", "Canada*", "2026-01-01", 5080),

    # Category 12A - Alloy merchant bars and light sections
    (58100, "Alloy merchant bars and light sections", "European Union", "2026-01-01", 29517),
    (58102, "Alloy merchant bars and light sections", "All others", "2026-01-01", 4246),

    # Category 12B - Non-alloy merchant bars and light sections
    (58103, "Non-alloy merchant bars and light sections", "European Union", "2026-01-01", 35458),
    (58104, "Non-alloy merchant bars and light sections", "Turkey", "2026-01-01", 13346),
    (58105, "Non-alloy merchant bars and light sections", "All others", "2026-01-01", 7585),

    # Category 13 - Rebars (UPDATED for 2026 - 058019 -> 058020)
    (58018, "Rebars", "European Union", "2026-01-01", 75521),
    (58866, "Rebars", "Turkey", "2026-01-01", 35304),
    (58020, "Rebars", "All others", "2026-01-01", 23514),  # NEW: was 58019
    (58131, "Rebars", "Egypt*", "2026-01-01", 3195),       # NEW individual country
    (58136, "Rebars", "Vietnam*", "2026-01-01", 4703),    # NEW individual country
    (58130, "Rebars", "Algeria*", "2026-01-01", 4703),    # NEW individual country
    (58133, "Rebars", "New Zealand*", "2026-01-01", 4703), # NEW individual country
    (58134, "Rebars", "Norway*", "2026-01-01", 4703),     # NEW individual country
    (58020, "Rebars", "Taiwan*", "2026-01-01", 4703),     # Uses same order as All others

    # Category 16 - Non-alloy and Other Alloy Wire Rod
    (58026, "Non-alloy and Other Alloy Wire Rod", "European Union", "2026-01-01", 75021),
    (58027, "Non-alloy and Other Alloy Wire Rod", "All others", "2026-01-01", 3275),

    # Category 17 - Angles, Shapes and Sections
    (58028, "Angles, Shapes and Sections of Iron or Non-alloy Steel", "European Union", "2026-01-01", 170816),
    (58029, "Angles, Shapes and Sections of Iron or Non-alloy Steel", "All others", "2026-01-01", 17753),

    # Category 20 - Gas pipes
    (58032, "Gas pipes", "European Union", "2026-01-01", 4797),
    (58033, "Gas pipes", "All others", "2026-01-01", 141),

    # Category 21 - Hollow sections
    (58034, "Hollow sections", "European Union", "2026-01-01", 7079),
    (58912, "Hollow sections", "India", "2026-01-01", 3640),
    (58916, "Hollow sections", "Turkey", "2026-01-01", 15728),
    (58035, "Hollow sections", "All others", "2026-01-01", 740),

    # Category 25A - Large welded tubes
    (58091, "Large welded tubes - 25A", "European Union", "2026-01-01", 11332),
    (58967, "Large welded tubes - 25A", "Turkey", "2026-01-01", 37335),
    (58036, "Large welded tubes - 25A", "All others", "2026-01-01", 3445),
    (58091, "Large welded tubes - 25A", "European Union", "2026-01-01", 6333),
    (58108, "Large welded tubes - 25A", "Japan", "2026-01-01", 8276),
    (58095, "Large welded tubes - 25A", "Republic of Korea", "2026-01-01", 1265),
    (58036, "Large welded tubes - 25A", "All others", "2026-01-01", 2232),

    # Category 25B - Large welded tubes
    (58038, "Large welded tubes - 25B", "All others", "2026-01-01", 16290),
    (58037, "Large welded tubes - 25B", "European Union", "2026-01-01", 2051),
    (58109, "Large welded tubes - 25B", "Japan", "2026-01-01", 4682),
    (58974, "Large welded tubes - 25B", "Republic of Korea", "2026-01-01", 4938),

    # Category 26 - Other Welded Pipes
    (58041, "Other Welded Pipes", "All others", "2026-01-01", 22820),
    (58039, "Other Welded Pipes", "European Union", "2026-01-01", 5860),
    (58949, "Other Welded Pipes", "China", "2026-01-01", 11140),
    (58947, "Other Welded Pipes", "Turkey", "2026-01-01", 15335),
    (58948, "Other Welded Pipes", "United Arab Emirates", "2026-01-01", 10118),
]


def create_uk_quota_file(output_path):
    """Create UK quota file with same format as EU file"""

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1: Base URL (matching EU format - URL in column B)
    ws.cell(1, 2).value = "https://www.trade-tariff.service.gov.uk/quota_search?order_number="

    # Rows 2-4: Empty (matching EU format)

    # Row 5: Headers (now including Template Quota Limit for UK scraper compatibility)
    headers = ["Order Number", "Quota Category", "Country", "Current Quarter", "Template Quota Limit", "URL"]
    for col, header in enumerate(headers, 1):
        ws.cell(5, col).value = header

    # Row 6+: Data with formula
    for row_idx, (order_num, category, country, quarter, template_limit) in enumerate(uk_quotas, 6):
        ws.cell(row_idx, 1).value = order_num
        ws.cell(row_idx, 2).value = category
        ws.cell(row_idx, 3).value = country
        ws.cell(row_idx, 4).value = quarter
        ws.cell(row_idx, 5).value = template_limit
        # Formula to build URL (pads order number to 6 digits)
        ws.cell(row_idx, 6).value = '=CONCATENATE($B$1,TEXT(Table2[[#This Row],[Order Number]],"000000"))'

    # Create table
    table_ref = f"A5:F{5 + len(uk_quotas)}"
    table = Table(displayName="Table2", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Adjust column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 70

    # Save
    wb.save(output_path)
    return len(uk_quotas)


if __name__ == "__main__":
    import os

    # Get project root (2 levels up from this script)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(script_dir))

    output_path = os.path.join(project_root, "data", "input", "uk_quota_urls.xlsx")

    count = create_uk_quota_file(output_path)
    print(f"Created: {output_path}")
    print(f"Total quota entries: {count}")
