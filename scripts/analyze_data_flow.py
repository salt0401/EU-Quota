# -*- coding: utf-8 -*-
"""
Comprehensive analysis of all Excel files to understand data flow
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, "logs")

def analyze_excel_structure(filepath, name):
    """Analyze Excel file structure including sheets, formatting, and interactive elements"""
    print(f"\n{'='*80}")
    print(f"ANALYZING: {name}")
    print(f"File: {filepath}")
    print(f"{'='*80}")

    result = {"name": name, "filepath": filepath, "sheets": []}

    try:
        wb = load_workbook(filepath, data_only=False)
        wb_data = load_workbook(filepath, data_only=True)

        print(f"\nSheet names: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_data = wb_data[sheet_name]

            sheet_info = {
                "name": sheet_name,
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "merged_cells": [str(m) for m in ws.merged_cells.ranges],
                "frozen_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
                "has_autofilter": ws.auto_filter.ref if ws.auto_filter.ref else None,
                "data_validations": [],
                "hyperlinks": [],
                "formulas": [],
                "headers": [],
                "sample_data": []
            }

            print(f"\n--- Sheet: {sheet_name} ---")
            print(f"  Dimensions: {ws.dimensions}")
            print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
            print(f"  Merged cells: {len(ws.merged_cells.ranges)}")
            print(f"  Frozen panes: {ws.freeze_panes}")
            print(f"  Auto filter: {ws.auto_filter.ref}")

            # Data validations (dropdowns, etc)
            for dv in ws.data_validations.dataValidation:
                dv_info = {
                    "type": dv.type,
                    "cells": str(dv.cells),
                    "formula1": str(dv.formula1) if dv.formula1 else None,
                    "formula2": str(dv.formula2) if dv.formula2 else None
                }
                sheet_info["data_validations"].append(dv_info)
                print(f"  Data validation: {dv.type} at {dv.cells}")

            # Hyperlinks
            for link in ws.hyperlinks:
                link_info = {
                    "cell": link.ref,
                    "target": link.target
                }
                sheet_info["hyperlinks"].append(link_info)
            if ws.hyperlinks:
                print(f"  Hyperlinks: {len(ws.hyperlinks)}")

            # Get headers (first non-empty row or known header row)
            header_row = 1
            for row in range(1, min(10, ws.max_row + 1)):
                values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
                if any(v for v in values):
                    header_row = row
                    break

            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(header_row, col).value
                if val:
                    headers.append({"col": col, "letter": get_column_letter(col), "value": str(val)})
            sheet_info["headers"] = headers
            print(f"  Headers (row {header_row}): {[h['value'] for h in headers[:10]]}{'...' if len(headers) > 10 else ''}")

            # Sample data rows
            for row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
                row_data = {}
                for col in range(1, min(ws.max_column + 1, 20)):
                    cell = ws.cell(row, col)
                    cell_data = ws_data.cell(row, col)
                    val = cell.value
                    computed_val = cell_data.value

                    # Check for formulas
                    if val and isinstance(val, str) and val.startswith('='):
                        sheet_info["formulas"].append({
                            "cell": f"{get_column_letter(col)}{row}",
                            "formula": val,
                            "computed": computed_val
                        })
                        row_data[get_column_letter(col)] = f"[FORMULA: {val[:50]}...]" if len(val) > 50 else f"[FORMULA: {val}]"
                    else:
                        row_data[get_column_letter(col)] = str(val) if val is not None else ""

                sheet_info["sample_data"].append(row_data)

            if sheet_info["formulas"]:
                print(f"  Formulas found: {len(sheet_info['formulas'])}")
                for f in sheet_info["formulas"][:3]:
                    print(f"    {f['cell']}: {f['formula'][:60]}...")

            result["sheets"].append(sheet_info)

        wb.close()
        wb_data.close()

    except Exception as e:
        print(f"Error: {e}")
        result["error"] = str(e)

    return result


def export_excel_to_csv(filepath, output_prefix):
    """Export all sheets from Excel to CSV for easy viewing"""
    try:
        xl = pd.ExcelFile(filepath)
        for sheet_name in xl.sheet_names:
            # Try different header rows
            for header_row in [0, 1, 2, 3, 4, 5]:
                try:
                    df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row)
                    if len(df.columns) > 0 and not df.columns[0].startswith('Unnamed'):
                        break
                except:
                    continue

            safe_name = sheet_name.replace(' ', '_').replace('/', '_')
            csv_path = os.path.join(OUTPUT_DIR, f"{output_prefix}_{safe_name}.csv")
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"  Exported: {csv_path}")
    except Exception as e:
        print(f"  Error exporting: {e}")


def main():
    files_to_analyze = [
        ("data/input/quota_urls.xlsx", "INPUT_quota_urls"),
        ("data/output/eu_quota_report_20260115.xlsx", "OUTPUT_raw_report"),
        ("data/output/eu_quota_report_20260115_customer.xlsx", "OUTPUT_customer_report"),
        ("data/snapshots/snapshot_20260115_151343.xlsx", "SNAPSHOT_data"),
        ("templates/meps_customer_template.xlsx", "TEMPLATE_meps_customer"),
    ]

    all_results = {}

    for rel_path, name in files_to_analyze:
        filepath = os.path.join(BASE_DIR, rel_path)
        if os.path.exists(filepath):
            result = analyze_excel_structure(filepath, name)
            all_results[name] = result

            # Export to CSV
            print(f"\n  Exporting to CSV...")
            export_excel_to_csv(filepath, name)
        else:
            print(f"\nFile not found: {filepath}")

    # Save comprehensive analysis
    output_file = os.path.join(OUTPUT_DIR, "data_flow_analysis.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n\nComplete analysis saved to: {output_file}")


if __name__ == "__main__":
    main()
