# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
import os

output_file = r"c:\Users\lyen\Downloads\EU Quota\analysis.txt"
template_file = r"c:\Users\lyen\Downloads\EU Quota\MEPS European Steel Review Quota Update - Dec 2025.xlsx"
snapshot_file = r"c:\Users\lyen\Downloads\EU Quota\data\snapshots\snapshot_20260115_151343.xlsx"

with open(output_file, 'w', encoding='utf-8') as f:
    f.write("=== TEMPLATE FILE ===\n")
    wb1 = load_workbook(template_file)
    f.write(f"Sheets: {wb1.sheetnames}\n")
    for sn in wb1.sheetnames:
        ws = wb1[sn]
        f.write(f"\nSheet: {sn}\n")
        f.write(f"Dimensions: {ws.dimensions}, Max: {ws.max_row} x {ws.max_column}\n")
        if ws.merged_cells.ranges:
            f.write(f"Merged: {[str(r) for r in ws.merged_cells.ranges]}\n")
        f.write(f"Frozen: {ws.freeze_panes}\n")
        # Column widths
        cw = {k: v.width for k, v in ws.column_dimensions.items() if v.width}
        f.write(f"Col widths: {cw}\n")
        # Headers
        headers = [ws.cell(1, c).value for c in range(1, min(ws.max_column+1, 25))]
        f.write(f"Headers: {headers}\n")
        # Sample row 2
        row2 = [ws.cell(2, c).value for c in range(1, min(ws.max_column+1, 25))]
        f.write(f"Row2: {row2}\n")
        # Sample row 3
        row3 = [ws.cell(3, c).value for c in range(1, min(ws.max_column+1, 25))]
        f.write(f"Row3: {row3}\n")
        # Check formatting of header cell
        cell = ws.cell(1, 1)
        f.write(f"Header font: bold={cell.font.bold}, size={cell.font.size}\n")
        if cell.fill.start_color.type == 'rgb':
            f.write(f"Header fill: {cell.fill.start_color.rgb}\n")
    
    f.write("\n\n=== SNAPSHOT FILE ===\n")
    wb2 = load_workbook(snapshot_file)
    f.write(f"Sheets: {wb2.sheetnames}\n")
    for sn in wb2.sheetnames:
        ws = wb2[sn]
        f.write(f"\nSheet: {sn}\n")
        f.write(f"Dimensions: {ws.dimensions}, Max: {ws.max_row} x {ws.max_column}\n")
        headers = [ws.cell(1, c).value for c in range(1, min(ws.max_column+1, 25))]
        f.write(f"Headers: {headers}\n")
        row2 = [ws.cell(2, c).value for c in range(1, min(ws.max_column+1, 25))]
        f.write(f"Row2: {row2}\n")
        row3 = [ws.cell(3, c).value for c in range(1, min(ws.max_column+1, 25))]
        f.write(f"Row3: {row3}\n")
    
    f.write("\n\n=== PANDAS COMPARISON ===\n")
    # Template
    for sn in wb1.sheetnames:
        df = pd.read_excel(template_file, sheet_name=sn)
        f.write(f"\nTemplate sheet '{sn}': shape={df.shape}\n")
        f.write(f"Columns: {list(df.columns)}\n")
        f.write(f"Dtypes:\n{df.dtypes.to_string()}\n")
    
    # Snapshot
    for sn in wb2.sheetnames:
        df = pd.read_excel(snapshot_file, sheet_name=sn)
        f.write(f"\nSnapshot sheet '{sn}': shape={df.shape}\n")
        f.write(f"Columns: {list(df.columns)}\n")
        f.write(f"Dtypes:\n{df.dtypes.to_string()}\n")

print("Analysis complete. Output saved to:", output_file)
