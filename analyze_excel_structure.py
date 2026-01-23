# -*- coding: utf-8 -*-
"""
Analyze the structure of both Excel files to understand what formatting needs to be replicated.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border
import os

def analyze_excel_file(filepath, name):
    """Analyze an Excel file's structure and formatting"""
    print(f"\n{'='*80}")
    print(f"ANALYZING: {name}")
    print(f"File: {filepath}")
    print(f"{'='*80}")
    
    if not os.path.exists(filepath):
        print(f"❌ File not found: {filepath}")
        return
    
    # Load with openpyxl for detailed formatting info
    wb = load_workbook(filepath)
    
    print(f"\n📋 Sheet names: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: '{sheet_name}' ---")
        print(f"   Dimensions: {ws.dimensions}")
        print(f"   Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        # Check for merged cells
        if ws.merged_cells.ranges:
            print(f"   Merged cells: {list(ws.merged_cells.ranges)}")
        
        # Check frozen panes
        if ws.freeze_panes:
            print(f"   Frozen panes: {ws.freeze_panes}")
        
        # Check column widths
        print(f"\n   Column widths:")
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.width:
                print(f"      {col_letter}: {col_dim.width}")
        
        # Check row heights
        custom_heights = [(row_num, row_dim.height) for row_num, row_dim in ws.row_dimensions.items() if row_dim.height]
        if custom_heights[:5]:
            print(f"\n   Custom row heights (first 5): {custom_heights[:5]}")
        
        # Analyze header row (row 1)
        print(f"\n   Header row (row 1):")
        for col in range(1, min(ws.max_column + 1, 20)):  # Limit to first 20 columns
            cell = ws.cell(row=1, column=col)
            if cell.value:
                fill_color = cell.fill.start_color.rgb if cell.fill.start_color.type == 'rgb' else 'default'
                font_color = cell.font.color.rgb if cell.font.color and cell.font.color.type == 'rgb' else 'default'
                print(f"      Col {col}: '{cell.value}'")
                print(f"         Font: bold={cell.font.bold}, size={cell.font.size}, color={font_color}")
                print(f"         Fill: {fill_color}")
                print(f"         Alignment: h={cell.alignment.horizontal}, v={cell.alignment.vertical}")
        
        # Sample data rows (rows 2-5)
        print(f"\n   Sample data (rows 2-5, first 10 columns):")
        for row in range(2, min(6, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                row_data.append(str(cell.value)[:30] if cell.value else "")
            print(f"      Row {row}: {row_data}")
        
        # Check for conditional formatting
        if ws.conditional_formatting._cf_rules:
            print(f"\n   Conditional formatting rules: {len(ws.conditional_formatting._cf_rules)}")
            for range_str, rules in list(ws.conditional_formatting._cf_rules.items())[:3]:
                print(f"      Range: {range_str}, Rules: {len(rules)}")
        
        # Check for data validation
        if ws.data_validations.dataValidation:
            print(f"\n   Data validations: {len(ws.data_validations.dataValidation)}")
    
    # Also read with pandas for data comparison
    print(f"\n📊 Pandas DataFrames:")
    for sheet_name in wb.sheetnames:
        df = pd.read_excel(filepath, sheet_name=sheet_name)
        print(f"\n   Sheet '{sheet_name}':")
        print(f"      Shape: {df.shape}")
        print(f"      Columns: {list(df.columns)}")
        print(f"      Data types:\n{df.dtypes.to_string()}")
    
    return wb

# Paths
template_file = r"c:\Users\lyen\Downloads\EU Quota\MEPS European Steel Review Quota Update - Dec 2025.xlsx"
snapshot_file = r"c:\Users\lyen\Downloads\EU Quota\data\snapshots\snapshot_20260115_151343.xlsx"

# Analyze both files
template_wb = analyze_excel_file(template_file, "TEMPLATE (MEPS European Steel Review)")
snapshot_wb = analyze_excel_file(snapshot_file, "SNAPSHOT (Scraped Data)")

print("\n" + "="*80)
print("ANALYSIS COMPLETE")
print("="*80)
