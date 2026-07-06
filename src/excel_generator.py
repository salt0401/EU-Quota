# -*- coding: utf-8 -*-
"""
EU Quota Excel Generator
Creates MEPS-style Excel reports by modifying template XML directly

Strategy: Copy the template file and modify XML directly using zipfile,
preserving all slicers, drawings, and interactive elements that openpyxl strips.

NOTE: This module uses string-based XML manipulation (not ElementTree) to preserve
Excel's complex namespace prefixes which are required for slicers to work properly.
"""

import os
import shutil
import zipfile
import tempfile
import re
import pandas as pd
from datetime import date, datetime
from typing import Optional, Tuple

from .data_processor import extract_period_info, prepare_customer_data, get_quota_summary
from .utils import get_templates_folder


def prepare_uk_customer_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare UK DataFrame for customer-facing MEPS report

    Selects and renames columns to match MEPS template format.
    UK data uses different column names than EU data.

    Args:
        df: Processed DataFrame with UK quota metrics

    Returns:
        pd.DataFrame: Customer-ready DataFrame with proper column names
    """
    if df is None or df.empty:
        return pd.DataFrame()

    # Exclude rows whose scrape failed or where the API had no data for the
    # order number: they carry only template-derived figures and would render
    # as untouched quotas (0% allocated / 100% remaining) in the customer
    # sheet. Mirrors the EU exclusion in data_processor.prepare_customer_data;
    # the rows remain in the UK raw-data output.
    if 'scrape_status' in df.columns:
        failed = df['scrape_status'] == 'failed'
        if failed.any():
            print(f"  Warning: excluding {int(failed.sum())} failed UK quota(s) "
                  f"from the customer report: "
                  f"{sorted(df.loc[failed, 'input_order_number'].astype(str))}")
            df = df[~failed]
            if df.empty:
                return pd.DataFrame()

    # Column mapping: internal_name -> display_name
    # UK scraper uses _tonnes suffix for already-converted values
    column_mapping = {
        'input_quota_category': 'Quota Category',
        'input_country': 'Country',
        'quota_limit_tonnes': 'Quota Limit (Tonnes)',
        'quota_allocated_tonnes': 'Quota Allocated (Tonnes)',
        'pct_allocated': '% Quota Allocated',
        'balance_remaining_tonnes': 'Balance Remaining (Tonnes)',
        'pct_remaining': '% Balance Remaining',
    }

    # Check if required columns exist, if not try to calculate
    if 'quota_limit_tonnes' not in df.columns and 'opening_balance_kg' in df.columns:
        df = df.copy()
        df['quota_limit_tonnes'] = df['opening_balance_kg'].apply(
            lambda x: x / 1000 if pd.notna(x) else 0
        )
    if 'quota_allocated_tonnes' not in df.columns and 'allocated_kg' in df.columns:
        if 'quota_limit_tonnes' not in df.columns:
            df = df.copy()
        df['quota_allocated_tonnes'] = df['allocated_kg'].apply(
            lambda x: x / 1000 if pd.notna(x) else 0
        )
    if 'balance_remaining_tonnes' not in df.columns and 'current_balance_kg' in df.columns:
        if 'quota_limit_tonnes' not in df.columns:
            df = df.copy()
        df['balance_remaining_tonnes'] = df['current_balance_kg'].apply(
            lambda x: x / 1000 if pd.notna(x) else 0
        )

    # Calculate percentages if not present
    if 'pct_allocated' not in df.columns and 'quota_limit_tonnes' in df.columns and 'quota_allocated_tonnes' in df.columns:
        df['pct_allocated'] = df.apply(
            lambda r: r['quota_allocated_tonnes'] / r['quota_limit_tonnes']
            if pd.notna(r['quota_limit_tonnes']) and r['quota_limit_tonnes'] > 0 else 0,
            axis=1
        )
    if 'pct_remaining' not in df.columns and 'quota_limit_tonnes' in df.columns and 'balance_remaining_tonnes' in df.columns:
        df['pct_remaining'] = df.apply(
            lambda r: r['balance_remaining_tonnes'] / r['quota_limit_tonnes']
            if pd.notna(r['quota_limit_tonnes']) and r['quota_limit_tonnes'] > 0 else 0,
            axis=1
        )

    # Select columns that exist
    available_cols = [c for c in column_mapping.keys() if c in df.columns]

    if not available_cols:
        return pd.DataFrame()

    result = df[available_cols].copy()

    # Rename columns
    result = result.rename(columns={k: v for k, v in column_mapping.items() if k in result.columns})

    # Round numeric columns
    numeric_cols = ['Quota Limit (Tonnes)', 'Quota Allocated (Tonnes)', 'Balance Remaining (Tonnes)']
    for col in numeric_cols:
        if col in result.columns:
            result[col] = result[col].round(0)

    # Ensure percentages are in correct format (0-1 range)
    pct_cols = ['% Quota Allocated', '% Balance Remaining']
    for col in pct_cols:
        if col in result.columns:
            result[col] = result[col].round(4)

    return result


def generate_meps_report(
    df: pd.DataFrame,
    output_path: str,
    period_display: Optional[str] = None,
    latest_data_date: Optional[str] = None,
    uk_df: Optional[pd.DataFrame] = None
) -> str:
    """
    Generate MEPS-style Excel report by copying template and updating XML directly

    Args:
        df: Processed DataFrame with EU quota metrics
        output_path: Full path for output file
        period_display: Optional custom period string
        latest_data_date: Optional custom latest data date
        uk_df: Optional processed DataFrame with UK quota metrics

    Returns:
        str: Path to generated file
    """
    # Extract period info if not provided
    if period_display is None or latest_data_date is None:
        auto_period, auto_latest, quarter, year = extract_period_info(df)
        if period_display is None:
            period_display = auto_period
        if latest_data_date is None:
            latest_data_date = auto_latest

    # Prepare customer data
    customer_df = prepare_customer_data(df)

    # Prepare UK customer data if provided
    uk_customer_df = None
    if uk_df is not None and not uk_df.empty:
        uk_customer_df = prepare_uk_customer_data(uk_df)

    # Get template path
    template_path = os.path.join(get_templates_folder(), "meps_customer_template.xlsx")

    if os.path.exists(template_path):
        # Use XML-based update to preserve slicers
        return _update_template_xml(template_path, output_path, customer_df,
                                    period_display, latest_data_date, uk_customer_df)
    else:
        # Fallback: generate from scratch using openpyxl (no slicers)
        return _generate_from_scratch(output_path, customer_df,
                                      period_display, latest_data_date, uk_customer_df)


def _update_template_xml(
    template_path: str,
    output_path: str,
    df: pd.DataFrame,
    period_display: str,
    latest_data: str,
    uk_df: Optional[pd.DataFrame] = None
) -> str:
    """
    Update template using direct XML manipulation to preserve slicers

    This approach:
    1. Extracts template xlsx to temp directory
    2. Modifies sharedStrings.xml (for date text updates)
    3. Modifies sheet2.xml (for EU data)
    4. Modifies sheet3.xml (for UK data)
    5. Repackages the xlsx preserving all other files (slicers, drawings, etc.)
    """
    # Create a temporary directory for extraction
    with tempfile.TemporaryDirectory() as temp_dir:
        # Extract the template xlsx (it's a ZIP file)
        with zipfile.ZipFile(template_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        # Update sharedStrings.xml with new date strings
        shared_strings_path = os.path.join(temp_dir, 'xl', 'sharedStrings.xml')
        if os.path.exists(shared_strings_path):
            _update_shared_strings(shared_strings_path, period_display, latest_data)

        # The Instructions sheet mirrors the banners via formulas; their CACHED
        # values live in sheet1.xml and are what LibreOffice/previewers/pandas
        # display (they don't recalculate on load). Patch the caches too.
        sheet1_path = os.path.join(temp_dir, 'xl', 'worksheets', 'sheet1.xml')
        if os.path.exists(sheet1_path):
            _update_shared_strings(sheet1_path, period_display, latest_data)

        # Belt and braces: force a full recalculation when Excel opens the file
        workbook_path = os.path.join(temp_dir, 'xl', 'workbook.xml')
        if os.path.exists(workbook_path):
            with open(workbook_path, 'r', encoding='utf-8') as f:
                wb_content = f.read()
            if 'fullCalcOnLoad' not in wb_content:
                wb_content = re.sub(r'<calcPr\s+([^/>]*?)\s*/>',
                                    r'<calcPr \1 fullCalcOnLoad="1"/>',
                                    wb_content)
                with open(workbook_path, 'w', encoding='utf-8') as f:
                    f.write(wb_content)

        # Calculate last row for table/dimension updates. Never let a table
        # shrink to its header row alone (ref="A15:G15"): Excel treats a
        # zero-data-row table as corrupt and refuses to open the workbook.
        # An empty df gets a placeholder row 16 instead.
        last_row = 15 + max(1, len(df))

        # Update sheet2.xml (European Union) with new data
        sheet2_path = os.path.join(temp_dir, 'xl', 'worksheets', 'sheet2.xml')
        if os.path.exists(sheet2_path):
            _update_eu_sheet_xml(sheet2_path, df)

        # Update table1.xml to match new data range
        table1_path = os.path.join(temp_dir, 'xl', 'tables', 'table1.xml')
        if os.path.exists(table1_path):
            _update_table_xml(table1_path, last_row)

        # Update sheet3.xml (United Kingdom) with UK data or placeholder
        sheet3_path = os.path.join(temp_dir, 'xl', 'worksheets', 'sheet3.xml')
        if os.path.exists(sheet3_path):
            _update_uk_sheet_xml(sheet3_path, uk_df)

        # Update table2.xml for UK data range if UK data exists
        if uk_df is not None and not uk_df.empty:
            uk_last_row = 15 + len(uk_df)
            table2_path = os.path.join(temp_dir, 'xl', 'tables', 'table2.xml')
            if os.path.exists(table2_path):
                _update_table_xml(table2_path, uk_last_row)

        # Package to a temp file first, then move to output
        temp_output = os.path.join(temp_dir, 'output.xlsx')
        _repackage_xlsx_to_file(temp_dir, temp_output)

        # Try to write to output path, handle if locked
        try:
            if os.path.exists(output_path):
                os.remove(output_path)
            shutil.move(temp_output, output_path)
        except PermissionError:
            # File is locked (probably open in Excel), use alternative filename
            base, ext = os.path.splitext(output_path)
            alt_output = f"{base}_new{ext}"
            print(f"  Warning: {os.path.basename(output_path)} is locked, saving as {os.path.basename(alt_output)}")
            shutil.move(temp_output, alt_output)
            return alt_output

    return output_path


def _update_shared_strings(filepath: str, period_display: str, latest_data: str):
    """Update date strings in sharedStrings.xml"""

    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Update the Instructions title (e.g., "MEPS Quota Update - December 2025" -> current month/year)
    current_month_year = date.today().strftime("%B %Y")
    title_pattern = r'MEPS Quota Update - [^<]+'
    new_title = f'MEPS Quota Update - {current_month_year}'
    content = re.sub(title_pattern, new_title, content)

    # Update the period string (matches pattern like "Current quota period: XX-XXX-XXXX to XX-XXX-XXXX")
    period_pattern = r'Current quota period: [^<]+'
    new_period = f'Current quota period: {period_display}'
    content = re.sub(period_pattern, new_period, content)

    # Update the latest data string
    latest_pattern = r'Latest available data: [^<]+'
    new_latest = f'Latest available data: {latest_data}'
    content = re.sub(latest_pattern, new_latest, content)

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(content)


def _update_eu_sheet_xml(filepath: str, df: pd.DataFrame):
    """
    Update European Union sheet with new data rows using string manipulation.
    This preserves all XML namespaces exactly as they are in the template.
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find the sheetData section
    sheet_data_start = content.find('<sheetData>')
    sheet_data_end = content.find('</sheetData>')

    if sheet_data_start == -1 or sheet_data_end == -1:
        print("Warning: Could not find sheetData in sheet2.xml")
        return

    # Extract the sheetData content
    sheet_data_content = content[sheet_data_start:sheet_data_end + len('</sheetData>')]

    # Find where row 15 ends (the header row) - keep rows 1-15
    # Look for the last occurrence of </row> before row 16 starts
    row_15_end = sheet_data_content.find('<row r="16"')
    if row_15_end == -1:
        # Try to find row 15's closing tag
        row_15_pattern = re.search(r'<row r="15"[^>]*>.*?</row>', sheet_data_content, re.DOTALL)
        if row_15_pattern:
            row_15_end = row_15_pattern.end()
        else:
            print("Warning: Could not find row 15 in sheet2.xml")
            return

    # Keep the header portion (rows 1-15)
    header_rows = sheet_data_content[len('<sheetData>'):row_15_end]

    # Generate new data rows as XML strings. If every scrape failed the
    # customer df is empty — write a placeholder row so the sheet (and the
    # A15:G16 table ref) stays valid for Excel, mirroring the UK sheet.
    if len(df) > 0:
        new_data_rows = []
        for idx, row_data in enumerate(df.values):
            row_num = 16 + idx
            row_xml = _create_data_row_xml(row_num, row_data)
            new_data_rows.append(row_xml)
        data_rows_xml = ''.join(new_data_rows)
    else:
        data_rows_xml = ('<row r="16" spans="1:7" ht="17.45">'
                         '<c r="A16" s="21" t="inlineStr"><is>'
                         '<t>EU data not available for this run</t></is></c></row>')

    # Build new sheetData section
    new_sheet_data = '<sheetData>' + header_rows + data_rows_xml + '</sheetData>'

    # Replace sheetData in the content
    new_content = content[:sheet_data_start] + new_sheet_data + content[sheet_data_end + len('</sheetData>'):]

    # Update dimension ref (min 16: the placeholder row)
    last_row = 15 + max(1, len(df))
    new_content = re.sub(
        r'<dimension ref="A1:G\d+"',
        f'<dimension ref="A1:G{last_row}"',
        new_content
    )

    # Update protected range
    new_content = _update_protected_range(new_content, last_row)

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(new_content)


def _create_data_row_xml(row_num: int, data: tuple) -> str:
    """
    Create an XML row string for data.
    Uses inline strings for text and proper formatting for numbers.
    Style indices: s="21" for text, s="22" for numbers, s="24" for percentages
    """
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    cells = []

    for col_idx, value in enumerate(data):
        cell_ref = f'{col_letters[col_idx]}{row_num}'

        if col_idx < 2:  # Text columns (Quota Category, Country)
            # Escape XML special characters
            text_value = str(value) if value else ''
            text_value = text_value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            cells.append(f'<c r="{cell_ref}" s="21" t="inlineStr"><is><t>{text_value}</t></is></c>')
        elif col_idx in [4, 6]:  # Percentage columns (% Quota Allocated, % Balance Remaining)
            # Both EU (prepare_customer_data) and UK (prepare_uk_customer_data)
            # supply 0-1 fractions; write as-is for the '0%' cell format.
            if pd.isna(value):
                num_value = '0'
            else:
                num_value = str(float(value))
            cells.append(f'<c r="{cell_ref}" s="24"><v>{num_value}</v></c>')
        else:  # Numeric columns (Quota Limit, Quota Allocated, Balance Remaining)
            if pd.isna(value):
                num_value = '0'
            else:
                num_value = str(value)
            cells.append(f'<c r="{cell_ref}" s="22"><v>{num_value}</v></c>')

    return f'<row r="{row_num}" spans="1:7" ht="17.45">{"".join(cells)}</row>'


def _update_uk_sheet_xml(filepath: str, uk_df: Optional[pd.DataFrame] = None):
    """
    Update UK sheet with data or placeholder.
    Uses string manipulation to preserve all XML namespaces.

    Args:
        filepath: Path to sheet3.xml
        uk_df: DataFrame with UK quota data (optional)
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find the sheetData section
    sheet_data_start = content.find('<sheetData>')
    sheet_data_end = content.find('</sheetData>')

    if sheet_data_start == -1 or sheet_data_end == -1:
        return

    # Extract the sheetData content
    sheet_data_content = content[sheet_data_start:sheet_data_end + len('</sheetData>')]

    # Find where row 15 ends (the header row) - keep rows 1-15
    row_15_end = sheet_data_content.find('<row r="16"')
    if row_15_end == -1:
        row_15_pattern = re.search(r'<row r="15"[^>]*>.*?</row>', sheet_data_content, re.DOTALL)
        if row_15_pattern:
            row_15_end = row_15_pattern.end()
        else:
            return

    # Keep the header portion (rows 1-15)
    header_rows = sheet_data_content[len('<sheetData>'):row_15_end]

    # Generate data rows or placeholder
    if uk_df is not None and not uk_df.empty:
        # Generate actual UK data rows
        new_data_rows = []
        for idx, row_data in enumerate(uk_df.values):
            row_num = 16 + idx
            row_xml = _create_data_row_xml(row_num, row_data)
            new_data_rows.append(row_xml)
        data_rows_xml = ''.join(new_data_rows)

        # Update dimension ref
        last_row = 15 + len(uk_df)
    else:
        # Create placeholder row
        data_rows_xml = '<row r="16" spans="1:7" ht="17.45"><c r="A16" s="21" t="inlineStr"><is><t>UK data not yet available</t></is></c></row>'
        last_row = 16

    # Build new sheetData section
    new_sheet_data = '<sheetData>' + header_rows + data_rows_xml + '</sheetData>'

    # Replace sheetData in the content
    new_content = content[:sheet_data_start] + new_sheet_data + content[sheet_data_end + len('</sheetData>'):]

    # Update dimension ref for UK sheet
    new_content = re.sub(
        r'<dimension ref="A1:G\d+"',
        f'<dimension ref="A1:G{last_row}"',
        new_content
    )

    # Update protected range
    new_content = _update_protected_range(new_content, last_row)

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(new_content)


def _update_table_xml(filepath: str, last_row: int):
    """
    Update table definition to match new data range.
    Updates ref attribute and autoFilter ref.
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Update table ref (e.g., ref="A15:G204" -> ref="A15:G{last_row}")
    content = re.sub(
        r'ref="A15:G\d+"',
        f'ref="A15:G{last_row}"',
        content
    )

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(content)


def _update_protected_range(content: str, last_row: int) -> str:
    """Update protected range in sheet XML to match new data range."""
    return re.sub(
        r'<protectedRange sqref="A15:G\d+"',
        f'<protectedRange sqref="A15:G{last_row}"',
        content
    )


def _repackage_xlsx_to_file(source_dir: str, output_path: str):
    """Repackage extracted files back into xlsx"""

    # Skip the output file itself if it exists in source_dir
    skip_file = os.path.basename(output_path)

    # Create new zip with xlsx extension
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(source_dir):
            for file in files:
                if file == skip_file:
                    continue
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, source_dir)
                zipf.write(file_path, arcname)


def _generate_from_scratch(
    output_path: str,
    df: pd.DataFrame,
    period_display: str,
    latest_data: str,
    uk_df: Optional[pd.DataFrame] = None
) -> str:
    """
    Generate report from scratch using openpyxl (fallback if template not available)
    Note: This won't have slicers
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    MEPS_BLUE = "16477C"
    HEADER_FILL = PatternFill(start_color=MEPS_BLUE, end_color=MEPS_BLUE, fill_type="solid")
    HEADER_FONT = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
    DATA_FONT = Font(name="Segoe UI", size=11)
    TITLE_FONT = Font(name="Kalinga", bold=True, size=26, color="FFFFFF")
    SUBTITLE_FONT = Font(name="Kalinga", size=12, color="FFFFFF")

    wb = Workbook()

    # Create sheets
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_eu = wb.create_sheet("European Union")
    ws_uk = wb.create_sheet("United Kingdom")

    # Instructions sheet title (dynamic month/year)
    current_month_year = date.today().strftime("%B %Y")
    ws_inst['A1'] = f"MEPS Quota Update - {current_month_year}"
    ws_inst['A1'].font = TITLE_FONT

    # Build EU Data sheet
    # Set header background
    for row in range(1, 15):
        for col in range(1, 8):
            ws_eu.cell(row=row, column=col).fill = HEADER_FILL

    ws_eu['A1'] = "MEPS European Union Quota Update"
    ws_eu['A1'].font = TITLE_FONT

    ws_eu['A2'] = f"Current quota period: {period_display}"
    ws_eu['A3'] = f"Latest available data: {latest_data}"
    ws_eu['A2'].font = SUBTITLE_FONT
    ws_eu['A3'].font = SUBTITLE_FONT

    # Data table starts at row 15
    start_row = 15
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_eu.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write data
    for row_idx, row_data in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_eu.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal='right')
            if col_idx in (5, 7):  # percentage columns hold 0-1 fractions
                cell.number_format = '0%'

    # Set column widths
    col_widths = [64, 40, 20, 20, 15, 20, 15]
    for col_idx, width in enumerate(col_widths, 1):
        ws_eu.column_dimensions[get_column_letter(col_idx)].width = width

    ws_eu.freeze_panes = f'A{start_row + 1}'

    # UK sheet
    for row in range(1, 15):
        for col in range(1, 8):
            ws_uk.cell(row=row, column=col).fill = HEADER_FILL

    ws_uk['A1'] = "MEPS United Kingdom Quota Update"
    ws_uk['A1'].font = TITLE_FONT
    ws_uk['A2'] = f"Current quota period: {period_display}"
    ws_uk['A3'] = f"Latest available data: {latest_data}"
    ws_uk['A2'].font = SUBTITLE_FONT
    ws_uk['A3'].font = SUBTITLE_FONT

    if uk_df is not None and not uk_df.empty:
        # Write UK headers
        uk_headers = list(uk_df.columns)
        for col_idx, header in enumerate(uk_headers, 1):
            cell = ws_uk.cell(row=start_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write UK data
        for row_idx, row_data in enumerate(uk_df.values, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_uk.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                if col_idx >= 3:
                    cell.alignment = Alignment(horizontal='right')
                if col_idx in (5, 7):  # percentage columns hold 0-1 fractions
                    cell.number_format = '0%'

        for col_idx, width in enumerate(col_widths, 1):
            ws_uk.column_dimensions[get_column_letter(col_idx)].width = width

        ws_uk.freeze_panes = f'A{start_row + 1}'
    else:
        ws_uk['A15'] = "UK data not yet available"

    wb.save(output_path)
    return output_path


def save_raw_data(df: pd.DataFrame, output_path: str) -> str:
    """Save raw scraped data to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    MEPS_BLUE = "16477C"
    HEADER_FILL = PatternFill(start_color=MEPS_BLUE, end_color=MEPS_BLUE, fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"

    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL

    # Write data
    for row_idx, row_data in enumerate(df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust widths
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(len(str(header)) + 5, 30)

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    return output_path


def save_snapshot(df: pd.DataFrame, output_folder: str) -> str:
    """Save timestamped snapshot"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"snapshot_{timestamp}.xlsx"
    output_path = os.path.join(output_folder, filename)

    df_copy = df.copy()
    df_copy['snapshot_timestamp'] = datetime.now().isoformat()

    save_raw_data(df_copy, output_path)
    return output_path
