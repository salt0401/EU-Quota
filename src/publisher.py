# -*- coding: utf-8 -*-
"""
Published-data writer for the daily GitHub Actions run.

Maintains data/published/ — the small, stable set of files that the
downloader program fetches from GitHub over public raw URLs:

    MEPS_Quota_Update_latest.xlsx   latest customer report (copied)
    quota_history_<YEAR>.csv        append-only daily history, one row per
                                    quota per day, one file per calendar
                                    year (canonical analysis data)
    Quota_History_<YEAR>.xlsx       the same history as a formatted workbook
    metadata.json                   freshness stamp + run summary + the list
                                    of files the downloader should fetch

This is a long-lived project: the history is split by year so no single
file grows forever. The history append is idempotent: re-running on the
same date replaces that date's rows instead of duplicating them.
"""

import csv
import glob
import io
import json
import os
import shutil
from datetime import datetime, date, timezone
from typing import Optional, List, Dict

import pandas as pd

HISTORY_COLUMNS = [
    'date', 'region', 'order_number', 'quota_category', 'country',
    'quota_limit_t', 'quota_allocated_t', 'pct_allocated',
    'balance_remaining_t', 'pct_remaining',
    'awaiting_allocation_t', 'validity_start', 'validity_end', 'status',
    'scrape_status',
]

PUBLISHED_FILES = {
    'report': 'MEPS_Quota_Update_latest.xlsx',
    'metadata': 'metadata.json',
}


def history_csv_name(year) -> str:
    return f'quota_history_{year}.csv'


def history_xlsx_name(year) -> str:
    return f'Quota_History_{year}.xlsx'


def _iso_date(value, fmt) -> str:
    """Parse a date string in the given format to ISO YYYY-MM-DD, else ''."""
    if not value or (isinstance(value, float) and pd.isna(value)):
        return ''
    try:
        return datetime.strptime(str(value).strip(), fmt).strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        return ''


def _round(value, ndigits=2):
    try:
        if value is None or pd.isna(value):
            return None
        return round(float(value), ndigits)
    except (TypeError, ValueError):
        return None


def build_eu_history_rows(eu_df: pd.DataFrame, run_date: str) -> List[Dict]:
    """One history row per EU quota. Internal EU metrics are kg and 0-100 %."""
    rows = []
    if eu_df is None or eu_df.empty:
        return rows
    for _, r in eu_df.iterrows():
        country = r.get('input_country')
        if country is None or (isinstance(country, float) and pd.isna(country)) or str(country).strip() == '':
            country = r.get('origin', '')
        status = r.get('scrape_status')
        rows.append({
            'date': run_date,
            'region': 'EU',
            'order_number': str(r.get('input_order_number', '')),
            'quota_category': r.get('input_quota_category', ''),
            'country': country,
            'quota_limit_t': _round(r.get('quota_limit', 0) / 1000 if pd.notna(r.get('quota_limit')) else None),
            'quota_allocated_t': _round(r.get('quota_allocated', 0) / 1000 if pd.notna(r.get('quota_allocated')) else None),
            'pct_allocated': _round(r.get('pct_allocated')),
            'balance_remaining_t': _round(r.get('balance_remaining', 0) / 1000 if pd.notna(r.get('balance_remaining')) else None),
            'pct_remaining': _round(r.get('pct_remaining')),
            # awaiting allocation decides real availability early in a quarter
            # (pooled requests can exceed the whole quota before allocation)
            'awaiting_allocation_t': _round(r.get('awaiting_allocation', 0) / 1000 if pd.notna(r.get('awaiting_allocation')) else None),
            'validity_start': _iso_date(r.get('validity_start'), '%d-%m-%Y'),
            'validity_end': _iso_date(r.get('validity_end'), '%d-%m-%Y'),
            'status': '',  # TARIC has no status field (Critical is always No, Art. 5)
            'scrape_status': status if isinstance(status, str) and status else 'ok',
        })
    return rows


def build_uk_history_rows(uk_df: pd.DataFrame, run_date: str) -> List[Dict]:
    """One history row per UK quota. Internal UK metrics are tonnes and 0-1 %."""
    rows = []
    if uk_df is None or uk_df.empty:
        return rows
    for _, r in uk_df.iterrows():
        status = r.get('scrape_status')
        pct_alloc = r.get('pct_allocated')
        pct_rem = r.get('pct_remaining')
        # UK validity arrives as one string, e.g. '01 July 2026 to 30 September 2026'
        v_start = v_end = ''
        period = r.get('validity_period')
        if isinstance(period, str) and ' to ' in period:
            start_s, _, end_s = period.partition(' to ')
            v_start = _iso_date(start_s, '%d %B %Y')
            v_end = _iso_date(end_s, '%d %B %Y')
        quota_status = r.get('status')
        rows.append({
            'date': run_date,
            'region': 'UK',
            'order_number': str(r.get('input_order_number', '')),
            'quota_category': r.get('input_quota_category', ''),
            'country': r.get('input_country', ''),
            'quota_limit_t': _round(r.get('quota_limit_tonnes')),
            'quota_allocated_t': _round(r.get('quota_allocated_tonnes')),
            'pct_allocated': _round(pct_alloc * 100 if pd.notna(pct_alloc) else None),
            'balance_remaining_t': _round(r.get('balance_remaining_tonnes')),
            'pct_remaining': _round(pct_rem * 100 if pd.notna(pct_rem) else None),
            'awaiting_allocation_t': None,  # not a UK API concept
            'validity_start': v_start,
            'validity_end': v_end,
            'status': quota_status if isinstance(quota_status, str) else '',
            'scrape_status': status if isinstance(status, str) and status else 'ok',
        })
    return rows


def update_history_csv(history_path: str, new_rows: List[Dict], run_date: str) -> int:
    """Append today's rows to the history CSV, replacing any existing rows for
    the same date (idempotent re-runs). Returns the total row count.

    Replacement is per (date, region): a re-run that produced rows for only
    one region keeps the other region's existing rows for that date instead
    of silently deleting them.
    """
    new_regions = {r['region'] for r in new_rows}
    existing = []
    if os.path.exists(history_path):
        with io.open(history_path, 'r', encoding='utf-8-sig', newline='') as f:
            existing = [row for row in csv.DictReader(f)
                        if not (row.get('date') == run_date
                                and row.get('region') in new_regions)]

    all_rows = existing + new_rows
    all_rows.sort(key=lambda r: (r['date'], r['region'], r['order_number']))

    os.makedirs(os.path.dirname(history_path), exist_ok=True)
    with io.open(history_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=HISTORY_COLUMNS)
        writer.writeheader()
        for row in all_rows:
            writer.writerow({k: ('' if row.get(k) is None else row.get(k)) for k in HISTORY_COLUMNS})
    return len(all_rows)


def generate_history_xlsx(history_path: str, xlsx_path: str):
    """Render the history CSV as a workbook with one filterable sheet per region."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    MEPS_BLUE = "16477C"
    header_fill = PatternFill(start_color=MEPS_BLUE, end_color=MEPS_BLUE, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    with io.open(history_path, 'r', encoding='utf-8-sig', newline='') as f:
        rows = list(csv.DictReader(f))

    numeric_cols = {'quota_limit_t', 'quota_allocated_t', 'pct_allocated',
                    'balance_remaining_t', 'pct_remaining', 'awaiting_allocation_t'}

    wb = Workbook()
    wb.remove(wb.active)
    for region, sheet_name in [('EU', 'EU History'), ('UK', 'UK History')]:
        ws = wb.create_sheet(sheet_name)
        for c, h in enumerate(HISTORY_COLUMNS, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
        r_idx = 2
        for row in rows:
            if row.get('region') != region:
                continue
            for c, h in enumerate(HISTORY_COLUMNS, 1):
                v = row.get(h, '')
                if h in numeric_cols and v not in ('', None):
                    try:
                        v = float(v)
                    except ValueError:
                        pass
                ws.cell(row=r_idx, column=c, value=v)
            r_idx += 1
        last_col = get_column_letter(len(HISTORY_COLUMNS))
        ws.auto_filter.ref = f'A1:{last_col}{max(r_idx - 1, 1)}'
        ws.freeze_panes = 'A2'
        widths = [12, 8, 14, 52, 30, 14, 16, 13, 18, 13, 18, 13, 13, 12, 13]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    wb.save(xlsx_path)


def _check_publish_gates(eu_df: pd.DataFrame, eu_rows: List[Dict],
                         uk_rows: List[Dict], run_date: date):
    """Refuse to publish datasets that would mislead: a mostly-failed scrape,
    or EU data whose quota window already expired (stale StartDate — TARIC
    serves expired windows as successful scrapes)."""
    for label, rows in (('EU', eu_rows), ('UK', uk_rows)):
        if rows:
            failed = sum(1 for r in rows if r['scrape_status'] != 'ok')
            if failed / len(rows) > 0.5:
                raise RuntimeError(
                    f"Refusing to publish: {failed}/{len(rows)} {label} quotas "
                    f"failed to scrape — the source site is likely broken or "
                    f"the quota definitions no longer exist (regulation "
                    f"renewal?). Fix the scrape before publishing.")

    if eu_df is not None and not eu_df.empty and 'validity_end' in eu_df.columns:
        ends = eu_df['validity_end'].dropna().astype(str)
        if len(ends) > 0:
            try:
                modal_end = datetime.strptime(ends.mode().iloc[0], '%d-%m-%Y').date()
                if modal_end < run_date:
                    raise RuntimeError(
                        f"Refusing to publish: the dominant EU quota window "
                        f"ended {modal_end} (before run date {run_date}) — the "
                        f"input workbook's 'Current Quarter' StartDate is stale "
                        f"and TARIC returned a frozen expired window.")
            except ValueError:
                pass  # unparseable dates: don't block on the gate itself


def publish_data(
    eu_df: pd.DataFrame,
    uk_df: Optional[pd.DataFrame],
    report_path: str,
    publish_dir: str,
    run_date: Optional[date] = None,
    period_display: str = '',
) -> dict:
    """Write/update everything in data/published/. Returns a summary dict.

    Write order: canonical text data (csv, metadata) first, lock-prone xlsx
    artifacts last — a workbook left open in Excel locally must not leave the
    canonical data half-published.
    """
    if run_date is None:
        run_date = date.today()
    date_str = run_date.strftime('%Y-%m-%d')
    os.makedirs(publish_dir, exist_ok=True)

    eu_rows = build_eu_history_rows(eu_df, date_str)
    uk_rows = build_uk_history_rows(uk_df, date_str)
    _check_publish_gates(eu_df, eu_rows, uk_rows, run_date)

    # 1. history csv (one file per calendar year; idempotent per date+region)
    current_csv = history_csv_name(run_date.year)
    current_xlsx = history_xlsx_name(run_date.year)
    history_path = os.path.join(publish_dir, current_csv)
    total_rows = update_history_csv(history_path, eu_rows + uk_rows, date_str)

    # every year's history file present in the repo, plus the matching
    # workbook names (past years' workbooks stay frozen on the release)
    history_csvs = sorted(os.path.basename(p) for p in
                          glob.glob(os.path.join(publish_dir, 'quota_history_*.csv')))
    release_workbooks = [PUBLISHED_FILES['report']] + [
        history_xlsx_name(name[len('quota_history_'):-len('.csv')])
        for name in history_csvs]

    eu_failed = sum(1 for r in eu_rows if r['scrape_status'] != 'ok')
    uk_failed = sum(1 for r in uk_rows if r['scrape_status'] != 'ok')
    # 2. metadata for the downloader's freshness check + file manifest
    metadata = {
        'generated_utc': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'data_date': date_str,
        'quota_period': period_display,
        'eu_quotas': len(eu_rows),
        'uk_quotas': len(uk_rows),
        'eu_failed': eu_failed,
        'uk_failed': uk_failed,
        'history_rows': total_rows,          # rows in the CURRENT year's file
        'current_history_csv': current_csv,
        'history_csvs': history_csvs,        # fetched from raw (git)
        'release_workbooks': release_workbooks,  # fetched from the release
        'files': history_csvs + release_workbooks + [PUBLISHED_FILES['metadata']],
    }
    with io.open(os.path.join(publish_dir, PUBLISHED_FILES['metadata']), 'w', encoding='utf-8') as f:
        json.dump(metadata, f, indent=2)

    # 3. latest customer report + history workbook (lock-prone: a file open in
    # Excel locally must not crash the publish after the canonical data landed)
    try:
        shutil.copy2(report_path, os.path.join(publish_dir, PUBLISHED_FILES['report']))
    except PermissionError:
        print(f"  Warning: {PUBLISHED_FILES['report']} is locked (open in "
              f"Excel?) — skipped updating it; csv/metadata are current.")
    try:
        generate_history_xlsx(history_path, os.path.join(publish_dir, current_xlsx))
    except PermissionError:
        print(f"  Warning: {current_xlsx} is locked (open in "
              f"Excel?) — skipped regenerating it; {current_csv} is current.")

    print(f"  Published to {publish_dir}: report + history ({total_rows} rows, "
          f"{len(eu_rows)} EU / {len(uk_rows)} UK today)")
    return metadata
