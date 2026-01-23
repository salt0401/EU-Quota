# -*- coding: utf-8 -*-
"""
EU Quota Data Exporter
Export scraped data to Excel, CSV, and snapshot formats

Author: Data Intern @ MEPS International
"""

import os
import pandas as pd
from datetime import datetime
from typing import Optional


def export_to_excel(df: pd.DataFrame, 
                   filename: str = None,
                   output_dir: str = None,
                   sheet_name: str = 'EU Quotas') -> str:
    """
    Export DataFrame to Excel file with formatting
    
    Args:
        df: DataFrame to export
        filename: Output filename (auto-generated if not provided)
        output_dir: Output directory (uses current dir if not provided)
        sheet_name: Excel sheet name
        
    Returns:
        str: Path to saved file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("❌ openpyxl is required: pip install openpyxl")
        return None
    
    if df.empty:
        print("❌ No data to export")
        return None
    
    # Create output directory if needed
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    else:
        output_dir = '.'
    
    # Generate filename if not provided
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"eu_quota_report_{timestamp}.xlsx"
    
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    filepath = os.path.join(output_dir, filename)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel max sheet name length
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(filepath)
    print(f"💾 Saved Excel: {filepath}")
    print(f"   📊 Rows: {len(df)}")
    
    return filepath


def export_to_csv(df: pd.DataFrame,
                 filename: str = None,
                 output_dir: str = None) -> str:
    """
    Export DataFrame to CSV file (Power BI compatible)
    
    Args:
        df: DataFrame to export
        filename: Output filename
        output_dir: Output directory
        
    Returns:
        str: Path to saved file
    """
    if df.empty:
        print("❌ No data to export")
        return None
    
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    else:
        output_dir = '.'
    
    if not filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"eu_quota_data_{timestamp}.csv"
    
    if not filename.endswith('.csv'):
        filename += '.csv'
    
    filepath = os.path.join(output_dir, filename)
    
    df.to_csv(filepath, index=False, encoding='utf-8-sig')
    print(f"💾 Saved CSV: {filepath}")
    
    return filepath


def save_snapshot(df: pd.DataFrame,
                 snapshot_dir: str = None,
                 prefix: str = 'snapshot') -> str:
    """
    Save timestamped snapshot of quota data as Excel file
    
    Args:
        df: DataFrame to save
        snapshot_dir: Directory for snapshots
        prefix: Filename prefix
        
    Returns:
        str: Path to saved snapshot
    """
    if snapshot_dir is None:
        snapshot_dir = os.path.join('.', 'data', 'snapshots')
    
    os.makedirs(snapshot_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{prefix}_{timestamp}.xlsx"
    filepath = os.path.join(snapshot_dir, filename)
    
    # Add snapshot metadata
    df = df.copy()
    df['snapshot_timestamp'] = datetime.now().isoformat()
    
    # Save as Excel
    df.to_excel(filepath, index=False, engine='openpyxl')
    print(f"📸 Saved snapshot: {filepath}")
    
    return filepath


def load_latest_snapshot(snapshot_dir: str = None,
                        prefix: str = 'snapshot') -> Optional[pd.DataFrame]:
    """
    Load the most recent snapshot from the snapshot directory
    
    Args:
        snapshot_dir: Directory containing snapshots
        prefix: Filename prefix to filter
        
    Returns:
        pd.DataFrame or None if no snapshots found
    """
    if snapshot_dir is None:
        snapshot_dir = os.path.join('.', 'data', 'snapshots')
    
    if not os.path.exists(snapshot_dir):
        return None
    
    # Find all snapshot files
    files = [f for f in os.listdir(snapshot_dir) 
             if f.startswith(prefix) and f.endswith('.csv')]
    
    if not files:
        return None
    
    # Sort by name (timestamp in filename ensures correct order)
    files.sort(reverse=True)
    latest = files[0]
    
    filepath = os.path.join(snapshot_dir, latest)
    print(f"📂 Loading snapshot: {latest}")
    
    return pd.read_csv(filepath)
